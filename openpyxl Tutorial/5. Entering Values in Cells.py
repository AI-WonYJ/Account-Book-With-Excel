from openpyxl import Workbook
from random import *

wb = Workbook()  # 워크북을 하나 생성하고 wb라는 변수에 만든다.
ws = wb.active  # 활성화 된 시트를 ws 변수로 지정
ws.title = 'test sheet'  # 시트명 'test sheet'

# A1 셀부터 B3 셀까지 값을 넣어준다.
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3
ws['B1'] = 1
ws['B2'] = 2
ws['B3'] = 3

print(ws['A1'])  # A1 셀 정보 출력

print(ws['A2'].value)  # A2 셀 값 출력

c = ws.cell(column = 3, row = 1, value = 10)  # 셀의 위치로 값 지정
print(c)
print(c.value)



# A1 셀부터 J10 셀까지 1 ~ 100을 순서대로 넣는다.
i = 1
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(column = x, row = y, value = 1)
        i += 1

wb.save('cell_5_1.xlsx')
wb.close()



# A1 셀부터 J10 셀까지 1 ~ 100을 무작위로 넣는다.
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(column = x, row = y, value = randint(0, 100))
wb.save('cell_5_2.xlsx')
wb.close()
