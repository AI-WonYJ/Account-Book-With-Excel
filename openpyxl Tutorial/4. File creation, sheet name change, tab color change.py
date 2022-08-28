# 엑셀파일을 만들어서 저장
from openpyxl import Workbook

wb = Workbook()
ws = wb.active  # 활성화 된 시트를 담을 변수
ws.title = 'test_sheet_4'  # 시트의 이름 변경
wb.save('sample_4_1.xlsx')  # 현 파일 위치에 만들어진 엑셀파일을 'sample_4.xlsx'라는 이름으로 저장
wb.close()  # 워크북을 닫아준다.



# 엑셀 시트를 만든 다음 탭의 색깔 변경
from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()  # 시트 생성
ws.title = "mysheet"
ws.sheet_properties.tabColor = "ff0000"  # 탭 색상 변경
wb.save('sample_4_2.xlsx')
