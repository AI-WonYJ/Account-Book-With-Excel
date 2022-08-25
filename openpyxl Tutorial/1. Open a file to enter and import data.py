from openpyxl import load_workbook

wb = load_workbook('cell.xlsx')
ws = wb.active

for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(column = x, row = y).value, end = " " )
    print()