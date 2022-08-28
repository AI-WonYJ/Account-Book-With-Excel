import os  # 정해진 경로의 폴더내에 있는 엑셀 파일의 이름을 가져오기 위해 필요한 모듈
from openpyxl import Workbook  # 엑셀파일 생성
from openpyxl import load_workbook  # 폴더내의 엑셀파일 열기

path = "C:/Users/user/Desktop/Account Book/openpyxl Tutorial/3.test"

def data_input(ws, o_ws):  # 각 파일별로 데이터 넣기
    for row in ws.iter_rows():
        data = []  # 각 열(row)의 셀에 있는 데이터 값을 'data' 리스트에 추가
        for cell in row:
            data.append(cell.value)
        o_ws.append(data)  # 취합할 파일의 시트인 'o_ws'에 입력
        return o_ws

def file_input(path):  # 각 파일 가져오기
    files = os.listdir(path)  # 각 파일의 이름을 리스트 형태로 files'라는 변수에 저장
    o_wb = Workbook()  # 'o_wb'라는 워크북을 만든다.
    o_ws = o_wb.active  # 활성화 된 시트를 'o_ws'라는 변수로 정한다.
    for file in files:
        wb = load_workbook(path + '/' + file)  # 각 파일의 파일을 열어서 'wb'라는 변수에 저장
        ws = wb.active  # 'wb' 워크북의 활성화 된 시트를 'ws'라는 변수로 정한다.
        data_input(ws, o_ws)
    o_wb.save(path + '/' + 'Total.xlsx')  # 취합된 파일인 'o_wb'는 같은 폴더에 'Total.xlsx'라는 이름으로 저장

file_input(path)
