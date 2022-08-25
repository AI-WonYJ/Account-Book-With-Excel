# 엑셀파일 불러오기
from openpyxl import load_workbook  # 엑셀파일을 불러오기 위해서는 load_workbook 모듈이 필요

wb = load_workbook('openpyxl Tutorial\Test.xlsx')  # 엑셀파일의 경로를 지정하고 'wb'라는 변수에 입력
ws = wb.active  # 엑셀파일에서 활성화 된 시트를 'ws'라는 변수에 입력한다.

print(ws['A1'].value)  # 'A1' 셀의 값을 불러온다.
print(ws.cell(row = 3, column = 1).value)  # '시트.cell(row = 행번호, column = 열번호).value'
print(ws.cell(3, 1).value) # 'row' 와 'column'을 생략하고 행, 열에 대한 숫자값만 넣어도 됨



# 데이터 입력
for i in range(1, 100):  # B1 ~ B99 셀까지 글자를 입력
    ws.cell(i, 2).value = '주말에는 파이썬 공부를 해야지'
wb.save('openpyxl Tutorial\Test.xlsx')



# 시트 하나 추가
wb.create_sheet('Test')  # 시트 추가
ws_test = wb['Test']  # 새로만든 시트를 'ws_test'라는 변수에 저장
wb.save('openpyxl Tutorial\Test.xlsx')



# 데이터 복사
for row in ws.iter_rows():  # 'iter_rows': 해당 시트에서 한 행씩 데이터를 가져오는 함수
    data = []
    for cell in row:  # 한 행씩 가져온 데이터를 'data'라는 변수에 list형식으로 저장
        data.append(cell.value)
    ws_test.append(data)  # 저장된 data변수의 list를 새로만든 'Test' 시트에 추가
wb.save('openpyxl Tutorial\Test.xlsx')



# 값만 가져오기
for row in ws.values:  # ws.values를 사용하면 시트에서 값만 가지고 옴
    data = []
    for cell in row:
        data.append(cell)
    print(data)



# 각 행, 열의 최소값과 최대값을 지정하기
for row in ws.iter_rows(min_col = 2, max_row = 50):  # 최소열을 2로 지정, 최대행을 50으로 지정
    data = []
    for cell in row:
        data.append(cell.value)
    ws_test.append(data)
wb.save('openpyxl Tutorial\Test.xlsx')
