from openpyxl import load_workbook

wb = load_workbook('cell_5_2.xlsx')  # 가져올 파일을 wb 변수에 담는다.
ws = wb.active  # 활성화 된 시트를 ws 변수에 담는다.

# column과 row를 1부터 10까지의 좌표에 값을 가져온다.
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(column = x, row = y).value, end = " " )
    print()



# 셀의 개수 (행과 열의 개수)를 모를 때는 다음을 이용한다.
for x in range(1, ws.max_column + 1):
    for y in range(1, ws.max_row + 1):
        print(ws.cell(row = y, column = x).value, end = " ")
    print()



# append를 사용해 셀에 값을 입력
from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(['Num', 'Eng', 'Math'])

for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])
wb.save('cell_range_6_1.xlsx')



# 데이터 가져오기 (B 열 가져오기)
col_B = ws['B']
print(col_B)
for cell in col_B:
    print(cell.value, end = " ")



# 첫번째 row만 가져오기
row_title = ws[1]
for cell in row_title:
    print(cell.value)



# 2번째 줄에서 6번째 줄까지 가져오기
row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
    print()



# 2번째 줄에서 마지막 줄까지 가져오기(max_row)
from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2: ws.max_row]
for rows in row_range:
    for cell in rows:
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end = " ")
    print()
